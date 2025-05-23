import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil

SEPARATOR_FILL = "FFF2F2F2"
HEADER_FILL_DARK = 'FF003366'
HEADER_FILL='FF808080'
CALCULATION_FILL='FFFFFF99'
LINKED_ENTRY_FILL='FFFFCC00'


def is_separator_cell(cell):
    is_empty = cell.value is None or str(cell.value).strip() == ""
    fill = cell.fill
    has_separator_fill = (
        fill is not None and fill.bgColor.rgb == SEPARATOR_FILL
    )
    border = cell.border
    no_border = all(
        getattr(border, side).style is None for side in ['left', 'right', 'top', 'bottom']
    )
    return is_empty and has_separator_fill and no_border

def is_header_cell(cell):
    fill = cell.fill
    return (
        fill is not None and (fill.bgColor.rgb == HEADER_FILL_DARK)
    )
def is_entry_cell(cell):
    fill=cell.fill
    return(
        fill is not None and fill.bgColor.rgb == SEPARATOR_FILL
    )
def is_formula_cell(cell):
    fill=cell.fill
    return(
        fill is not None and (fill.bgColor.rgb == CALCULATION_FILL or fill.bgColor.rgb == LINKED_ENTRY_FILL) 
    )

# def is_merged(cell, merged_ranges):
#     for mrange in merged_ranges:
#         min_col, min_row, max_col, max_row = range_boundaries(mrange)
#         if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
#             return (min_row, max_row, min_col, max_col)
#     return None

def detect_tables_and_headers(sheet):
    #merged_ranges = [str(rng) for rng in sheet.merged_cells.ranges]
    tables = []
    in_table = False
    table_start = None

    for idx, row in enumerate(sheet.iter_rows(), start=1):
        is_separator = all(is_separator_cell(cell) for cell in row)
        if is_separator:
            if in_table:
                tables.append((table_start, idx - 1))
                in_table = False
        else:
            if not in_table:
                table_start = idx
                in_table = True
    if in_table:
        tables.append((table_start, sheet.max_row))

    table_infos = []
    for start, end in tables:
        header_rows = set()
        header_cols = set()

        for r in range(start, end + 1):
            row = sheet[r]
            for cell in row:
                if is_header_cell(cell):
                    header_rows.add(cell.row)
                    header_cols.add(cell.column)
            
        # Trying to find out column header range 
        table_infos.append({
            "table_range": (start, end),
            "header_row_ranges": sorted(header_rows),
            "header_col_indices": sorted(header_cols),
            "header_row_depth": max(header_rows)-min(header_rows)+1 if header_rows else 0,
            "header_col_depth": max(header_cols)-min(header_cols) if header_cols else 0
        })

    print("TABLES INFO EXTRACTED :: ",table_infos)
    return table_infos

def extract_table_as_df(sheet, table_info):
    start_row, end_row = table_info["table_range"]
    header_rows = table_info["header_row_ranges"]
    header_cols = table_info["header_col_indices"]
    #merged_ranges = [str(rng) for rng in sheet.merged_cells.ranges]
    print("Trying to extract table as df..")
    # Step 1: Extract raw data block
    data = []
    for r in range(start_row, end_row + 1):
        row_values = []
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if(is_header_cell(cell)):
                row_values.append(val)
            else:
                if(is_entry_cell(cell)):
                    row_values.append(val)
                elif (is_formula_cell(cell)):
                    formulaText=f"{val}"
                    row_values.append(formulaText[1:])
                else :
                    row_values.append('')
        data.append(row_values)

    # # Step 2: Convert to DataFrame
    df_raw = pd.DataFrame(data)
    
    ## Slice headers
    headerRows=df_raw.iloc[:table_info["header_row_depth"],table_info["header_col_depth"]:]
    headerCols=df_raw.iloc[table_info["header_row_depth"]:,:table_info["header_col_depth"]:]
    coreData=df_raw.iloc[table_info["header_row_depth"]:,table_info["header_col_depth"]:]
    print("Header Rows: ",headerRows)
    print("Header Cols :",headerCols)
    df=pd.DataFrame(coreData)

    if not headerCols.empty:
        df.index=pd.MultiIndex.from_frame(headerCols.reset_index(drop=True))
    if not headerRows.empty:
        df.columns=pd.MultiIndex.from_arrays(headerRows.values)
    return df


def export_tables_to_excel_with_spacing(dataframes, output_path, spacing=3):
    wb = Workbook()
    ws = wb.active
    ws.title = "ParsedTables"

    current_row = 1

    for df in dataframes:
        # Write DataFrame (including index)
        print(list(df.columns))
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=0):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=current_row + r_idx, column=c_idx, value=value)
        current_row += len(df) + 2 + spacing  # +2 to account for header and index

    wb.save(output_path)
    print(f"\n✅ All tables exported to: {output_path}")

# Output path
output_file = "parsed_tables.xlsx"



# Load workbook
file_path = "Test_Data_v1.xlsx"
clone_file_path= "Test_Data_v1_clone.xlsx"
shutil.copyfile(file_path,clone_file_path)
wb = load_workbook(file_path,data_only=False, read_only= True)
sheet = wb.active

# Detect tables
tables = detect_tables_and_headers(sheet)

#Extract DataFrames
dfs = []
for idx, table in enumerate(tables):
    df = pd.DataFrame(extract_table_as_df(sheet, table))
    print(f"\n--- DataFrame {idx + 1} ---")
    print(df.head())
    dfs.append(df)
# Export all DataFrames
export_tables_to_excel_with_spacing(dfs, output_file, spacing=3)
